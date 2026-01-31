from openpyxl import load_workbook
import re

REDACT = "████████"


# Salary / pay in any currency (rupee, dollar, euro, dinar, pound)
SALARY_PATTERNS = [
    r"\b(?:salary|sal|pay|wage|income)\s*[:\-]?\s*\d[\d,.]*\s*(?:lakh|lac|k|cr|million|mn|bn)?\b",
    r"[₹$€£]\s*\d[\d,.]*\s*(?:lakh|lac|k|cr|m|mn|bn)?\b",
    r"\b(?:rs\.?|inr|rupee|rupees)\s*[:\-]?\s*\d[\d,.]*\s*(?:lakh|lac|k|cr)?\b",
    r"\b(?:usd|\$|dollar|dollars)\s*[:\-]?\s*\d[\d,.]*\s*(?:k|m|mn|million|bn)?\b",
    r"\b(?:eur|€|euro|euros)\s*[:\-]?\s*\d[\d,.]*\s*(?:k|m|mn|million)?\b",
    r"\b(?:gbp|£|pound|pounds)\s*[:\-]?\s*\d[\d,.]*\s*(?:k|m|mn)?\b",
    r"\b(?:dinar|dinars|kwd|bhd|jod|iqd|aed|sar)\s*[:\-]?\s*\d[\d,.]*\s*(?:k|m)?\b",
]


def is_sensitive(value):
    """Redact only cells that look like emails, phones, long numbers, keywords, names, or salary/currency."""
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    patterns = [
        r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+",  # email
        r"\b\d{10}\b",                                       # phone
        r"\b\d{4,}\b",                                       # long numbers
        r"\b(Aadhaar|PAN|Account|ATM|Password|IFSC|SSN|CVV)\b",
        r"\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,2})\b",         # Title Case names
    ]
    for p in patterns:
        if re.search(p, text, re.IGNORECASE if "Aadhaar" in p else 0):
            return True
    for p in SALARY_PATTERNS:
        if re.search(p, text, re.IGNORECASE):
            return True
    if re.match(r"^[A-Z][a-z]{1,30}$", text):
        return True
    return False


def redact_excel(input_path, output_path):
    wb = load_workbook(input_path)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str):
                    if is_sensitive(cell.value):
                        cell.value = REDACT
                elif isinstance(cell.value, (int, float)):
                    # Redact long numbers (e.g. IDs, phones)
                    s = str(cell.value)
                    if re.search(r"\d{10}", s) or (len(s) >= 4 and s.isdigit()):
                        cell.value = REDACT

    wb.save(output_path)
